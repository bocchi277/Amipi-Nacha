"""
Spreadsheet parser for payment upload batches.

Parses QuickBooks Desktop exports (.xlsx) and standard tabular CSV/Excel files.
Returns valid payment items and explicit per-row errors for malformed entries.
"""
from __future__ import annotations

import io
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Optional

import openpyxl
import pandas as pd
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Vendor
from app.nacha.validation import validate_routing_checksum


@dataclass
class ParsedRowError:
    row_number: int
    raw_data: dict[str, Any]
    errors: list[str]


@dataclass
class ParsedPayment:
    vendor_name: str
    amount: Decimal
    id_number: str
    effective_date: date
    vendor_id: Optional[Any] = None
    routing_number: Optional[str] = None
    account_number: Optional[str] = None
    account_type: Optional[str] = "checking"
    notes: Optional[str] = None
    invoice_breakdown: Optional[list[dict[str, Any]]] = None



@dataclass
class ParseResult:
    valid_payments: list[ParsedPayment] = field(default_factory=list)
    errors: list[ParsedRowError] = field(default_factory=list)
    total_rows_parsed: int = 0


def _parse_amount(value: Any) -> Optional[Decimal]:
    """Parse string or numeric value into a Decimal amount."""
    if value is None:
        return None
    s = str(value).replace(",", "").replace("$", "").strip()
    if not s:
        return None
    try:
        d = Decimal(s)
        return d if d != Decimal("0") else None
    except (InvalidOperation, ValueError, TypeError):
        return None


def _parse_date(value: Any) -> Optional[date]:
    """Parse date from datetime, date object, or string formats."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None

    s = str(value).strip()
    # Common formats: YYYY-MM-DD, MM/DD/YYYY, YYMMDD, YYYYMMDD
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%y%m%d", "%Y%m%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


# NACHA Entry Detail "Individual Identification Number" field width (positions 40-54).
_ID_FIELD_WIDTH = 15


def _compress_invoices(invoice_list: list[str]) -> str:
    """
    Compress multiple invoice numbers into the 15-character NACHA ID field.

    The previous implementation joined with '/' and hard-truncated, so
    ``['UDI261954', 'UDI261965', 'UDI261955']`` became ``'UDI261954/UDI26'`` and two
    invoice numbers were silently lost -- the vendor could not tell which invoices the
    payment covered.

    Strategy (ported from the v7 prototype's ``compressInvoiceNums``):

    1. A single invoice is passed through unchanged.
    2. Otherwise factor out the longest common prefix and list only the differing
       suffixes, e.g. ``UDI261954 / UDI261965 / UDI261955`` -> ``UDI261954/65/55``,
       which keeps every invoice identifiable inside 15 characters.
    3. If the plain '/'-joined form already fits, use it.
    4. If nothing fits (unrelated long invoice numbers), keep the first invoice and
       append a ``+N`` marker so it is EXPLICIT that more invoices are covered rather
       than appearing to be a single-invoice payment. The complete list always remains
       available in the payment's ``invoice_breakdown`` JSON and on the remittance
       advice, so no data is lost -- only the fixed-width bank field is abbreviated.
    """
    cleaned = [str(n).strip() for n in invoice_list if str(n).strip()]
    if not cleaned:
        return "EPAY"

    # De-duplicate while preserving order (QB exports repeat invoice numbers).
    seen: set[str] = set()
    unique: list[str] = []
    for item in cleaned:
        if item not in seen:
            seen.add(item)
            unique.append(item)

    if len(unique) == 1:
        return unique[0][:_ID_FIELD_WIDTH]

    # 2. Common-prefix compression.
    prefix = os.path.commonprefix(unique)
    if len(prefix) >= 3:
        suffixes = [s[len(prefix):] for s in unique]
        if all(suffixes):
            candidate = prefix + "/".join(suffixes)
            if len(candidate) <= _ID_FIELD_WIDTH:
                return candidate

    # 3. Plain join when it fits.
    plain = "/".join(unique)
    if len(plain) <= _ID_FIELD_WIDTH:
        return plain

    # 4. Explicit "first + N more" marker.
    marker = f"+{len(unique) - 1}"
    return unique[0][: _ID_FIELD_WIDTH - len(marker)] + marker


async def parse_payment_spreadsheet(
    file_bytes: bytes,
    filename: str,
    db_session: AsyncSession,
    default_effective_date: Optional[date] = None,
) -> ParseResult:
    """
    Parse an uploaded spreadsheet (.xlsx, .xls, .csv) into payment entries.
    """
    result = ParseResult()

    # Pre-fetch all active vendors from database for matching
    res = await db_session.execute(select(Vendor).where(Vendor.is_active == True))
    db_vendors = res.scalars().all()
    vendor_map = {v.name.strip().upper(): v for v in db_vendors}

    # Detect if file is QuickBooks report format or standard tabular file
    ext = filename.lower().split(".")[-1]
    
    if ext in ("xlsx", "xls"):
        result = _parse_excel(file_bytes, vendor_map, default_effective_date)
    elif ext == "csv":
        result = _parse_csv(file_bytes, vendor_map, default_effective_date)
    else:
        result.errors.append(
            ParsedRowError(
                row_number=0,
                raw_data={},
                errors=[f"Unsupported file extension '.{ext}'. Supported formats: .xlsx, .xls, .csv"],
            )
        )

    return result


def _parse_excel(
    file_bytes: bytes,
    vendor_map: dict[str, Vendor],
    default_effective_date: Optional[date] = None,
) -> ParseResult:
    """Parse Excel workbook (.xlsx). Handles QB reports & standard tables."""
    result = ParseResult()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result.errors.append(
            ParsedRowError(
                row_number=0,
                raw_data={},
                errors=[f"Corrupted or invalid Excel file format: {e}"],
            )
        )
        return result

    # Prefer 'Sheet1' or first non-tips sheet
    sheet_name = wb.sheetnames[0]
    for sname in wb.sheetnames:
        if "sheet" in sname.lower() or "pay" in sname.lower():
            sheet_name = sname
            break
    ws = wb[sheet_name]

    # Check if this is a QuickBooks export format (Type, Num, Date, Name, Paid Amount)
    header_row_idx = None
    col_idx_map = {}

    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        headers_upper = [h.upper() for h in row_vals]
        
        if "TYPE" in headers_upper and ("NAME" in headers_upper or any("PAID" in h for h in headers_upper)):
            header_row_idx = r
            paid_col = None
            orig_col = None
            generic_amt_col = None
            for c, h in enumerate(headers_upper, 1):
                if h == "TYPE":
                    col_idx_map["type"] = c
                elif h == "NUM":
                    col_idx_map["num"] = c
                elif h == "DATE":
                    col_idx_map["date"] = c
                elif h == "NAME":
                    col_idx_map["name"] = c
                elif "PAID" in h:
                    paid_col = c
                elif "ORIGINAL" in h:
                    orig_col = c
                elif "AMOUNT" in h:
                    generic_amt_col = c

            col_idx_map["amount"] = paid_col or generic_amt_col or orig_col
            break
        
        # Standard table headers check
        if any(k in headers_upper for k in ("VENDOR", "VENDOR NAME", "PAYEE", "NAME")):
            header_row_idx = r
            paid_col = None
            orig_col = None
            generic_amt_col = None
            for c, h in enumerate(headers_upper, 1):
                if h in ("VENDOR", "VENDOR NAME", "PAYEE", "NAME"):
                    col_idx_map["name"] = c
                elif "PAID" in h:
                    paid_col = c
                elif "ORIGINAL" in h:
                    orig_col = c
                elif h in ("AMOUNT", "TRXN AMOUNT") or "AMOUNT" in h:
                    generic_amt_col = c
                elif h in ("NUM", "INVOICE", "ID NUMBER", "ID"):
                    col_idx_map["num"] = c
                elif h in ("DATE", "EFFECTIVE DATE"):
                    col_idx_map["date"] = c
                elif h in ("ROUTING", "ROUTING NUMBER", "ROUTING NUM"):
                    col_idx_map["routing"] = c
                elif h in ("ACCOUNT", "ACCOUNT NUMBER", "ACCT NUMBER"):
                    col_idx_map["account"] = c

            if paid_col or generic_amt_col or orig_col:
                col_idx_map["amount"] = paid_col or generic_amt_col or orig_col
            break

    if header_row_idx and "type" in col_idx_map:
        # QuickBooks report format parser
        return _parse_qb_excel(ws, header_row_idx, col_idx_map, vendor_map, default_effective_date)

    if header_row_idx and "name" in col_idx_map and "amount" in col_idx_map:
        # Standard tabular Excel format parser
        return _parse_tabular_excel(ws, header_row_idx, col_idx_map, vendor_map, default_effective_date)

    # Fallback: scan all rows as QuickBooks grouped format (Paid Amount default at col 14)
    return _parse_qb_excel(ws, 1, {"type": 2, "num": 4, "date": 6, "name": 8, "amount": 14}, vendor_map, default_effective_date)


def _parse_qb_excel(
    ws,
    header_row_idx: int,
    col_map: dict[str, int],
    vendor_map: dict[str, Vendor],
    default_effective_date: Optional[date] = None,
) -> ParseResult:
    """Parse QuickBooks grouped report format (Bill Pmt -Check / Bill / TOTAL rows)."""
    result = ParseResult()
    current_vendor = None
    current_invoices = []
    current_amount = Decimal("0.00")
    current_date = None

    type_col = col_map.get("type", 2)
    num_col = col_map.get("num", 4)
    date_col = col_map.get("date", 6)
    name_col = col_map.get("name", 8)
    amt_col = col_map.get("amount", 14)

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        first_val = str(row_vals[0] or "").strip().upper()

        # Handle TOTAL row closing a vendor block
        if first_val.startswith("TOTAL"):
            # Check TOTAL row for authoritative amount in designated amount column or rightmost cells
            total_row_amt = None
            if len(row_vals) >= amt_col:
                total_row_amt = _parse_amount(row_vals[amt_col - 1])
            if total_row_amt is None or total_row_amt == Decimal("0"):
                for c in range(len(row_vals) - 1, -1, -1):
                    parsed_a = _parse_amount(row_vals[c])
                    if parsed_a is not None and parsed_a != Decimal("0"):
                        total_row_amt = abs(parsed_a)
                        break

            final_amt = abs(current_amount)
            # If total_row_amt is available and is greater than parsed line sum or line sum is 0,
            # use the total row as authoritative reconciliation amount
            if total_row_amt is not None and total_row_amt > 0:
                if final_amt == Decimal("0") or total_row_amt > final_amt:
                    final_amt = total_row_amt

            if current_vendor:
                _process_qb_vendor_block(
                    r, current_vendor, final_amt, current_invoices, current_date,
                    vendor_map, default_effective_date, result
                )
            current_vendor = None
            current_invoices = []
            current_amount = Decimal("0.00")
            current_date = None
            continue

        val_type = row_vals[type_col - 1] if len(row_vals) >= type_col else None
        row_type = str(val_type).strip() if val_type is not None and str(val_type).strip().lower() != "none" else ""

        val_name = row_vals[name_col - 1] if len(row_vals) >= name_col else None
        row_name = str(val_name).strip() if val_name is not None and str(val_name).strip().lower() != "none" else ""

        val_num = row_vals[num_col - 1] if len(row_vals) >= num_col else None
        row_num = str(val_num).strip() if val_num is not None and str(val_num).strip().lower() != "none" else ""

        val_date = row_vals[date_col - 1] if len(row_vals) >= date_col else None
        row_date = _parse_date(val_date)

        val_amt = row_vals[amt_col - 1] if len(row_vals) >= amt_col else None
        row_amt = _parse_amount(val_amt)

        # Fallback to search rightmost columns if row_amt is None and row is not a header row
        if row_amt is None and row_type.lower() not in ("bill pmt -check", "check", "payment", "bill pmt", "ach"):
            for c in range(len(row_vals) - 1, -1, -1):
                parsed_a = _parse_amount(row_vals[c])
                if parsed_a is not None and parsed_a > Decimal("0"):
                    row_amt = parsed_a
                    break

        # If a new vendor block starts and we have accumulated sub-invoices/amounts from previous vendor, flush it
        is_new_block_start = (
            row_type.lower() in ("bill pmt -check", "check", "payment", "bill pmt", "ach")
            or (row_name and current_vendor and row_name.upper() != current_vendor.upper())
        )
        if is_new_block_start and current_vendor and (current_amount > 0 or len(current_invoices) > 0):
            _process_qb_vendor_block(
                r - 1, current_vendor, abs(current_amount), current_invoices, current_date,
                vendor_map, default_effective_date, result
            )
            current_vendor = None
            current_invoices = []
            current_amount = Decimal("0.00")
            current_date = None

        if row_name:
            current_vendor = row_name
        if row_type.lower() in ("bill pmt -check", "check", "payment", "bill pmt", "ach") and row_date:
            current_date = row_date
        elif current_date is None and row_date:
            current_date = row_date

        if row_type.lower() in ("bill", "bill pmt -check", "check", "payment"):
            is_header_row = (
                row_type.lower() != "bill"
                or row_num.upper() in ("ACH", "CHECK", "EFT", "PMT", "PAYMENT", "BILL PMT -CHECK", "BILL PMT")
            )
            if row_num and not is_header_row:
                sub_amt = float(abs(row_amt)) if row_amt is not None else None
                sub_date = row_date.isoformat() if row_date else None
                current_invoices.append({
                    "invoice_number": row_num,
                    "amount": sub_amt,
                    "invoice_date": sub_date,
                })
            if row_amt is not None and row_type.lower() == "bill":
                current_amount += abs(row_amt)
        elif current_vendor and row_amt is not None and abs(row_amt) > 0:
            # Continuation / Split line within active vendor block (e.g. split expense row with blank Type/Num)
            current_amount += abs(row_amt)
            if current_invoices:
                prev_amt = current_invoices[-1].get("amount") or 0.0
                current_invoices[-1]["amount"] = round(prev_amt + float(abs(row_amt)), 2)
            elif row_num:
                sub_amt = float(abs(row_amt))
                sub_date = row_date.isoformat() if row_date else None
                current_invoices.append({
                    "invoice_number": row_num,
                    "amount": sub_amt,
                    "invoice_date": sub_date,
                })

    # Process trailing block if file ends without explicit TOTAL row
    if current_vendor and current_amount > 0:
        _process_qb_vendor_block(
            ws.max_row, current_vendor, abs(current_amount), current_invoices, current_date,
            vendor_map, default_effective_date, result
        )

    return result


def _process_qb_vendor_block(
    row_idx: int,
    vendor_name: str,
    amount: Decimal,
    invoices: list[Any],
    entry_date: Optional[date],
    vendor_map: dict[str, Vendor],
    default_effective_date: Optional[date],
    result: ParseResult,
) -> None:
    """Validate and record a parsed QuickBooks vendor payment block."""
    result.total_rows_parsed += 1
    raw_info = {"row": row_idx, "vendor": vendor_name, "amount": str(amount), "invoices": invoices}
    row_errors = []

    if not vendor_name:
        row_errors.append("Vendor name is required.")
    if amount <= 0:
        row_errors.append(f"Amount must be > 0, got {amount}.")

    eff_date = entry_date or default_effective_date or date.today()

    v_upper = vendor_name.strip().upper()
    v_obj = vendor_map.get(v_upper)
    if not v_obj:
        v_clean = re.sub(r'[^A-Z0-9]', '', v_upper)
        # Check fuzzy / substring / normalized match
        for db_name, db_v in vendor_map.items():
            db_clean = re.sub(r'[^A-Z0-9]', '', db_name.strip().upper())
            if (
                (len(db_clean) >= 4 and db_clean in v_clean)
                or (len(v_clean) >= 4 and v_clean in db_clean)
                or (len(db_clean) >= 4 and v_clean.startswith(db_clean))
                or db_name in v_upper
            ):
                v_obj = db_v
                break


    if not v_obj:
        row_errors.append(f"Vendor '{vendor_name}' not found in database and no banking routing/account provided.")

    if row_errors:
        result.errors.append(ParsedRowError(row_number=row_idx, raw_data=raw_info, errors=row_errors))
    else:
        filtered_invoices = [
            inv for inv in invoices
            if isinstance(inv, dict) and inv.get("invoice_number", "").upper() not in ("ACH", "CHECK", "EFT", "PMT", "PAYMENT", "BILL PMT -CHECK", "BILL PMT")
        ]
        inv_nums = [inv["invoice_number"] for inv in filtered_invoices]
        id_ref = _compress_invoices(inv_nums)
        breakdown_list = filtered_invoices if len(filtered_invoices) > 0 else None

        result.valid_payments.append(
            ParsedPayment(
                vendor_name=v_obj.name,
                amount=amount,
                id_number=id_ref,
                effective_date=eff_date,
                vendor_id=v_obj.id,
                routing_number=v_obj.routing_number,
                account_number=v_obj.account_number,
                account_type=v_obj.account_type.value if hasattr(v_obj.account_type, "value") else str(v_obj.account_type),
                invoice_breakdown=breakdown_list,
            )
        )


def _parse_tabular_excel(
    ws,
    header_row_idx: int,
    col_map: dict[str, int],
    vendor_map: dict[str, Vendor],
    default_effective_date: Optional[date] = None,
) -> ParseResult:
    """Parse standard tabular Excel file."""
    result = ParseResult()

    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(row_vals):
            continue

        result.total_rows_parsed += 1
        name_val = str(row_vals[col_map["name"] - 1] or "").strip() if "name" in col_map else ""
        amt_val = _parse_amount(row_vals[col_map["amount"] - 1]) if "amount" in col_map else None
        num_val = str(row_vals[col_map["num"] - 1] or "").strip() if "num" in col_map else "EPAY"
        date_val = _parse_date(row_vals[col_map["date"] - 1]) if "date" in col_map else None
        routing_val = str(row_vals[col_map["routing"] - 1] or "").strip() if "routing" in col_map else None
        acct_val = str(row_vals[col_map["account"] - 1] or "").strip() if "account" in col_map else None

        raw_info = {"row": r, "name": name_val, "amount": str(amt_val), "num": num_val}
        row_errors = []

        if not name_val:
            row_errors.append("Vendor name is required.")
        if amt_val is None or amt_val <= 0:
            row_errors.append(f"Amount must be a positive number.")

        eff_date = date_val or default_effective_date or date.today()
        v_obj = vendor_map.get(name_val.upper())

        if not v_obj and not (routing_val and acct_val):
            row_errors.append(f"Vendor '{name_val}' not found in database and no routing/account provided in row.")
        elif routing_val:
            if len(routing_val) != 9 or not validate_routing_checksum(routing_val):
                row_errors.append(f"Routing number '{routing_val}' is invalid.")

        if row_errors:
            result.errors.append(ParsedRowError(row_number=r, raw_data=raw_info, errors=row_errors))
        else:
            result.valid_payments.append(
                ParsedPayment(
                    vendor_name=v_obj.name if v_obj else name_val[:22],
                    amount=amt_val,
                    id_number=(num_val or "EPAY")[:15],
                    effective_date=eff_date,
                    vendor_id=v_obj.id if v_obj else None,
                    routing_number=routing_val or (v_obj.routing_number if v_obj else ""),
                    account_number=acct_val or (v_obj.account_number if v_obj else ""),
                    account_type=(v_obj.account_type.value if hasattr(v_obj.account_type, "value") else str(v_obj.account_type)) if v_obj else "checking",
                )
            )

    return result


def _clean_cell(val: Any) -> str:
    """Clean pandas / excel cell values converting NaN and None to empty string."""
    if val is None or pd.isna(val):
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def _parse_csv(
    file_bytes: bytes,
    vendor_map: dict[str, Vendor],
    default_effective_date: Optional[date] = None,
) -> ParseResult:
    """Parse CSV payment spreadsheet."""
    result = ParseResult()
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
    except Exception as e:
        result.errors.append(ParsedRowError(row_number=0, raw_data={}, errors=[f"CSV parsing error: {e}"]))
        return result

    # Normalize column names with explicit priority (Paid Amount > Amount > Original Amount)
    col_renames = {}
    paid_col = None
    amt_col = None
    orig_col = None

    for col in df.columns:
        c_upper = str(col).strip().upper()
        if c_upper in ("VENDOR", "VENDOR NAME", "PAYEE", "NAME"):
            col_renames[col] = "name"
        elif "PAID" in c_upper:
            paid_col = col
        elif "ORIGINAL" in c_upper:
            orig_col = col
        elif c_upper in ("AMOUNT", "TRXN AMOUNT") or "AMOUNT" in c_upper:
            amt_col = col
        elif c_upper in ("NUM", "INVOICE", "ID NUMBER", "ID", "INVOICE NUMBER"):
            col_renames[col] = "num"
        elif c_upper in ("DATE", "EFFECTIVE DATE"):
            col_renames[col] = "date"
        elif c_upper in ("ROUTING", "ROUTING NUMBER", "ROUTING NUM"):
            col_renames[col] = "routing"
        elif c_upper in ("ACCOUNT", "ACCOUNT NUMBER", "ACCT NUMBER"):
            col_renames[col] = "account"

    amount_chosen = paid_col or amt_col or orig_col
    if amount_chosen:
        col_renames[amount_chosen] = "amount"

    df = df.rename(columns=col_renames)

    for idx, row in df.iterrows():
        r = idx + 2  # 1-indexed row number (row 1 is header)
        name_val = _clean_cell(row.get("name"))
        amt_val = _parse_amount(row.get("amount"))
        num_val = _clean_cell(row.get("num"))
        date_val = _parse_date(row.get("date"))
        routing_val = _clean_cell(row.get("routing")) if "routing" in row else None
        acct_val = _clean_cell(row.get("account")) if "account" in row else None

        raw_info = {"row": r, "name": name_val, "amount": str(row.get("amount")), "num": num_val}
        row_errors = []

        if not name_val:
            row_errors.append("Vendor name is required.")
        if amt_val is None or amt_val <= 0:
            row_errors.append(f"Amount must be a positive number, got '{row.get('amount')}'.")

        eff_date = date_val or default_effective_date or date.today()
        v_obj = vendor_map.get(name_val.upper())

        if not v_obj and not (routing_val and acct_val):
            row_errors.append(f"Vendor '{name_val}' not found in database and no routing/account provided.")
        elif routing_val:
            if len(routing_val) != 9 or not validate_routing_checksum(routing_val):
                row_errors.append(f"Routing number '{routing_val}' is invalid.")

        result.total_rows_parsed += 1
        if row_errors:
            result.errors.append(ParsedRowError(row_number=r, raw_data=raw_info, errors=row_errors))
        else:
            result.valid_payments.append(
                ParsedPayment(
                    vendor_name=v_obj.name if v_obj else name_val[:22],
                    amount=amt_val,
                    id_number=(num_val or "EPAY")[:15],
                    effective_date=eff_date,
                    vendor_id=v_obj.id if v_obj else None,
                    routing_number=routing_val or (v_obj.routing_number if v_obj else ""),
                    account_number=acct_val or (v_obj.account_number if v_obj else ""),
                    account_type=(v_obj.account_type.value if hasattr(v_obj.account_type, "value") else str(v_obj.account_type)) if v_obj else "checking",
                )
            )

    return result
